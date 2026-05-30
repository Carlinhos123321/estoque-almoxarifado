"""MovStok ERP - Main REST API blueprint.

Exposes /api/* endpoints for every ERP module:
  - /api/dashboard
  - /api/products       (CRUD + filters + pagination + low-stock)
  - /api/categories     (CRUD)
  - /api/suppliers      (CRUD)
  - /api/employees      (CRUD)
  - /api/entries        (CRUD - stock in, updates product qty)
  - /api/outputs        (CRUD - stock out, decrements product qty)
  - /api/stock          (snapshot + low-stock alerts)
  - /api/reports/*      (data + PDF/Excel exports)
  - /api/notifications  (list/mark-read)
  - /api/activities     (audit log)
  - /api/users          (admin CRUD)
  - /api/roles          (list)
  - /api/company        (get/update)
  - /api/legacy/*       (backward-compatible endpoints for the old API)
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta
from decimal import Decimal
from flask import Blueprint, abort, jsonify, request, send_file
from flask_login import current_user, login_required
from sqlalchemy import desc, func, or_

from ..extensions import db
from ..helpers import log_activity, paginate_query, require_permission
from ..models import (
    ActivityLog, Category, Company, Employee, Notification, Permission,
    Product, Role, StockEntry, StockLocation, StockMovement, StockOutput, Supplier, User,
)
from ..services.dashboard_service import DashboardService
from ..services.product_service import ProductService
from ..services.stock_service import StockService
from ..services.audit_service import AuditService

bp = Blueprint("api", __name__)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _company_id():
    return getattr(current_user, "company_id", None)


def _ok(payload=None, status=200):
    return jsonify(payload if payload is not None else {"ok": True}), status


def _err(msg, status=400, **extra):
    body = {"error": msg}
    body.update(extra)
    return jsonify(body), status


def _to_decimal(v, default="0"):
    try:
        return Decimal(str(v if v not in (None, "") else default))
    except Exception:
        return Decimal(default)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@bp.get("/dashboard")
@require_permission("dashboard.view", api=True)
def dashboard():
    """Endpoint profissional: apenas orquestra a resposta da camada de serviço."""
    service = DashboardService(company_id=_company_id())
    data = service.get_summary_metrics()
    return _ok(data)


# ---------------------------------------------------------------------------
# Products
# ---------------------------------------------------------------------------
@bp.get("/products")
@require_permission("products.view", api=True)
def products_list():
    service = ProductService(_company_id(), current_user.id)
    filters = {
        "search": request.args.get("search"),
        "category_id": request.args.get("category_id", type=int),
        "stock_status": request.args.get("stock_status"),
        "sort": request.args.get("sort", "name"),
        "dir": request.args.get("dir", "asc")
    }
    result = service.list_products(filters, page=request.args.get("page", 1, type=int))
    return _ok({
        "items": [p.to_dict() for p in result["items"]],
        "meta": result["meta"],
    })


@bp.get("/products/<int:pid>")
@require_permission("products.view", api=True)
def products_get(pid):
    p = Product.query.filter_by(company_id=_company_id(), id=pid).first_or_404()
    return _ok(p.to_dict())


@bp.post("/products")
@require_permission("products.create", api=True)
def products_create():
    data = request.get_json(silent=True) or {}
    sku = (data.get("sku") or "").strip()
    name = (data.get("name") or "").strip()
    if not sku or not name:
        return _err("SKU e Nome são obrigatórios.")
    cid = _company_id()
    if Product.query.filter_by(company_id=cid, sku=sku).first():
        return _err("Já existe um produto com este SKU.")
    p = Product(
        company_id=cid,
        category_id=data.get("category_id") or None,
        supplier_id=data.get("supplier_id") or None,
        location_id=data.get("location_id") or None,
        sku=sku, name=name,
        description=data.get("description"),
        unit=(data.get("unit") or "UN").upper()[:20],
        barcode=data.get("barcode"),
        cost_price=_to_decimal(data.get("cost_price")),
        sale_price=_to_decimal(data.get("sale_price")),
        stock_quantity=_to_decimal(data.get("stock_quantity")),
        min_stock=_to_decimal(data.get("min_stock")),
        max_stock=_to_decimal(data.get("max_stock")),
        status=data.get("status") or "active",
        image_url=data.get("image_url"),
        created_by_id=current_user.id,
    )
    db.session.add(p)
    db.session.commit()
    log_activity("create", "product", p.id, f"Produto criado: {p.sku} - {p.name}")
    return _ok(p.to_dict(), 201)


@bp.put("/products/<int:pid>")
@bp.patch("/products/<int:pid>")
@require_permission("products.edit", api=True)
def products_update(pid):
    p = Product.query.filter_by(company_id=_company_id(), id=pid).first_or_404()
    data = request.get_json(silent=True) or {}
    for field in ("name", "description", "unit", "barcode", "image_url", "status"):
        if field in data:
            setattr(p, field, data[field])
    for field in ("category_id", "supplier_id", "location_id"):
        if field in data:
            setattr(p, field, data[field] or None)
    for field in ("cost_price", "sale_price", "min_stock", "max_stock"):
        if field in data:
            setattr(p, field, _to_decimal(data[field]))
    if "sku" in data and data["sku"]:
        new_sku = data["sku"].strip()
        if new_sku != p.sku and Product.query.filter_by(company_id=p.company_id, sku=new_sku).first():
            return _err("Já existe um produto com este SKU.")
        p.sku = new_sku
    # stock_quantity can be edited only via manual adjustment
    if "stock_quantity" in data and current_user.can("stock.entry"):
        p.stock_quantity = _to_decimal(data["stock_quantity"])
    p.updated_by_id = current_user.id
    db.session.commit()
    AuditService.log("update", "product", p.id, f"Produto atualizado: {p.sku}", _company_id(), current_user.id)
    return _ok(p.to_dict())


@bp.delete("/products/<int:pid>")
@require_permission("products.delete", api=True)
def products_delete(pid):
    p = Product.query.filter_by(company_id=_company_id(), id=pid).first_or_404()
    sku = p.sku
    p.status = "inactive"  # soft delete
    db.session.commit()
    log_activity("delete", "product", pid, f"Produto desativado: {sku}")
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Categories
# ---------------------------------------------------------------------------
@bp.get("/categories")
@require_permission("categories.view", api=True)
def categories_list():
    cid = _company_id()
    q = Category.query.filter_by(company_id=cid)
    search = (request.args.get("search") or "").strip()
    if search:
        q = q.filter(Category.name.ilike(f"%{search}%"))
    q = q.order_by(Category.name.asc())
    return _ok({"items": [c.to_dict() for c in q.all()]})


@bp.post("/categories")
@require_permission("categories.manage", api=True)
def categories_create():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return _err("Nome é obrigatório.")
    cid = _company_id()
    if Category.query.filter_by(company_id=cid, name=name).first():
        return _err("Já existe uma categoria com este nome.")
    c = Category(company_id=cid, name=name,
                 description=data.get("description"),
                 color=data.get("color") or "#1d4ed8",
                 created_by_id=current_user.id)
    db.session.add(c)
    db.session.commit()
    log_activity("create", "category", c.id, f"Categoria: {c.name}")
    return _ok(c.to_dict(), 201)


@bp.put("/categories/<int:cid>")
@bp.patch("/categories/<int:cid>")
@require_permission("categories.manage", api=True)
def categories_update(cid):
    c = Category.query.filter_by(company_id=_company_id(), id=cid).first_or_404()
    data = request.get_json(silent=True) or {}
    for f in ("name", "description", "color"):
        if f in data:
            setattr(c, f, data[f])
    if "active" in data:
        c.active = bool(data["active"])
    c.updated_by_id = current_user.id
    db.session.commit()
    AuditService.log("update", "category", c.id, f"Categoria: {c.name}", _company_id(), current_user.id)
    return _ok(c.to_dict())


@bp.delete("/categories/<int:cid>")
@require_permission("categories.manage", api=True)
def categories_delete(cid):
    c = Category.query.filter_by(company_id=_company_id(), id=cid).first_or_404()
    if c.products:
        return _err("Categoria possui produtos vinculados.")
    db.session.delete(c)
    db.session.commit()
    log_activity("delete", "category", cid, "Categoria removida")
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Suppliers
# ---------------------------------------------------------------------------
@bp.get("/suppliers")
@require_permission("suppliers.view", api=True)
def suppliers_list():
    cid = _company_id()
    q = Supplier.query.filter_by(company_id=cid)
    search = (request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        q = q.filter(or_(Supplier.name.ilike(like), Supplier.cnpj.ilike(like),
                         Supplier.email.ilike(like)))
    q = q.order_by(Supplier.name.asc())
    paginated = request.args.get("paginated", "1") == "1"
    if paginated:
        result = paginate_query(q, default_per_page=25)
        return _ok({"items": [s.to_dict() for s in result["items"]], "meta": result["meta"]})
    return _ok({"items": [s.to_dict() for s in q.all()]})


@bp.post("/suppliers")
@require_permission("suppliers.manage", api=True)
def suppliers_create():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return _err("Nome é obrigatório.")
    s = Supplier(
        company_id=_company_id(), name=name,
        cnpj=data.get("cnpj"), email=data.get("email"), phone=data.get("phone"),
        contact_person=data.get("contact_person"), address=data.get("address"),
        city=data.get("city"), state=data.get("state"), notes=data.get("notes"),
        active=bool(data.get("active", True)), created_by_id=current_user.id,
    )
    db.session.add(s)
    db.session.commit()
    log_activity("create", "supplier", s.id, f"Fornecedor: {s.name}")
    return _ok(s.to_dict(), 201)


@bp.put("/suppliers/<int:sid>")
@bp.patch("/suppliers/<int:sid>")
@require_permission("suppliers.manage", api=True)
def suppliers_update(sid):
    s = Supplier.query.filter_by(company_id=_company_id(), id=sid).first_or_404()
    data = request.get_json(silent=True) or {}
    for f in ("name", "cnpj", "email", "phone", "contact_person",
              "address", "city", "state", "notes"):
        if f in data:
            setattr(s, f, data[f])
    if "active" in data:
        s.active = bool(data["active"])
    s.updated_by_id = current_user.id
    db.session.commit()
    AuditService.log("update", "supplier", s.id, f"Fornecedor: {s.name}", _company_id(), current_user.id)
    return _ok(s.to_dict())


@bp.delete("/suppliers/<int:sid>")
@require_permission("suppliers.manage", api=True)
def suppliers_delete(sid):
    s = Supplier.query.filter_by(company_id=_company_id(), id=sid).first_or_404()
    s.active = False
    db.session.commit()
    log_activity("delete", "supplier", sid, f"Fornecedor desativado: {s.name}")
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Employees
# ---------------------------------------------------------------------------
@bp.get("/employees")
@require_permission("employees.view", api=True)
def employees_list():
    cid = _company_id()
    q = Employee.query.filter_by(company_id=cid)
    search = (request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        q = q.filter(or_(Employee.name.ilike(like), Employee.enrollment.ilike(like),
                         Employee.cpf.ilike(like), Employee.department.ilike(like)))
    status = request.args.get("status")
    if status:
        q = q.filter(Employee.status == status)
    q = q.order_by(Employee.name.asc())
    paginated = request.args.get("paginated", "1") == "1"
    if paginated:
        result = paginate_query(q, default_per_page=25)
        return _ok({"items": [e.to_dict() for e in result["items"]], "meta": result["meta"]})
    return _ok({"items": [e.to_dict() for e in q.all()]})


def _next_enrollment(cid):
    last = (Employee.query.filter_by(company_id=cid)
            .order_by(Employee.id.desc()).first())
    if last and last.enrollment.isdigit():
        return str(int(last.enrollment) + 1).zfill(4)
    return "0001"


@bp.post("/employees")
@require_permission("employees.manage", api=True)
def employees_create():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return _err("Nome é obrigatório.")
    cid = _company_id()
    enrollment = (data.get("enrollment") or "").strip() or _next_enrollment(cid)
    if Employee.query.filter_by(enrollment=enrollment).first():
        return _err("Matrícula já existente.")
    e = Employee(
        company_id=cid, enrollment=enrollment, name=name,
        cpf=data.get("cpf"), email=data.get("email"), phone=data.get("phone"),
        department=data.get("department"), position=data.get("position"),
        status=data.get("status") or "active", notes=data.get("notes"),
        created_by_id=current_user.id,
    )
    db.session.add(e)
    db.session.commit()
    log_activity("create", "employee", e.id, f"Funcionário: {e.enrollment} - {e.name}")
    return _ok(e.to_dict(), 201)


@bp.put("/employees/<int:eid>")
@bp.patch("/employees/<int:eid>")
@require_permission("employees.manage", api=True)
def employees_update(eid):
    e = Employee.query.filter_by(company_id=_company_id(), id=eid).first_or_404()
    data = request.get_json(silent=True) or {}
    for f in ("name", "cpf", "email", "phone", "department",
              "position", "status", "notes"):
        if f in data:
            setattr(e, f, data[f])
    e.updated_by_id = current_user.id
    db.session.commit()
    AuditService.log("update", "employee", e.id, f"Funcionário: {e.name}", _company_id(), current_user.id)
    return _ok(e.to_dict())


@bp.delete("/employees/<int:eid>")
@require_permission("employees.manage", api=True)
def employees_delete(eid):
    e = Employee.query.filter_by(company_id=_company_id(), id=eid).first_or_404()
    e.status = "terminated"
    db.session.commit()
    log_activity("delete", "employee", eid, f"Funcionário desligado: {e.name}")
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Stock Entries
# ---------------------------------------------------------------------------
@bp.get("/entries")
@require_permission("stock.view", api=True)
def entries_list():
    cid = _company_id()
    q = StockEntry.query.filter_by(company_id=cid)
    pid = request.args.get("product_id", type=int)
    if pid:
        q = q.filter(StockEntry.product_id == pid)
    sid = request.args.get("supplier_id", type=int)
    if sid:
        q = q.filter(StockEntry.supplier_id == sid)
    date_from = request.args.get("date_from")
    if date_from:
        try:
            q = q.filter(StockEntry.entry_date >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    date_to = request.args.get("date_to")
    if date_to:
        try:
            q = q.filter(StockEntry.entry_date <= datetime.fromisoformat(date_to))
        except ValueError:
            pass
    q = q.order_by(StockEntry.entry_date.desc(), StockEntry.id.desc())
    result = paginate_query(q, default_per_page=25)
    return _ok({"items": [e.to_dict() for e in result["items"]],
                "meta": result["meta"]})


@bp.post("/entries")
@require_permission("stock.entry", api=True)
def entries_create():
    try:
        service = StockService(_company_id(), current_user.id)
        entry = service.register_entry(request.get_json(silent=True) or {})
        return _ok(entry.to_dict(), 201)
    except ValueError as e:
        return _err(str(e))


@bp.delete("/entries/<int:eid>")
@require_permission("stock.entry", api=True)
def entries_delete(eid):
    e = StockEntry.query.filter_by(company_id=_company_id(), id=eid).first_or_404()
    if e.status == "cancelled":
        return _err("Entrada já cancelada.")
    p = e.product # p.company_id is verified via entry relationship
    p.stock_quantity = max(Decimal("0"), p.stock_quantity - e.quantity)
    e.status = "cancelled"
    db.session.add(StockMovement(
        company_id=e.company_id,
        product_id=e.product_id,
        entry_id=e.id,
        movement_type="cancel",
        document=e.document,
        quantity=-(e.quantity or 0),
        unit_value=e.unit_cost,
        total_value=-(e.total_cost or 0),
        balance_after=p.stock_quantity if p else 0,
        reason="Cancelamento de entrada",
        movement_date=datetime.utcnow(),
        created_by_id=current_user.id,
    ))
    db.session.commit()
    AuditService.log("delete", "stock_entry", e.id, f"Cancelamento entrada {e.document}", _company_id(), current_user.id)
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Stock Outputs
# ---------------------------------------------------------------------------
@bp.get("/outputs")
@require_permission("stock.view", api=True)
def outputs_list():
    cid = _company_id()
    q = StockOutput.query.filter_by(company_id=cid)
    pid = request.args.get("product_id", type=int)
    if pid:
        q = q.filter(StockOutput.product_id == pid)
    emp = request.args.get("employee_id", type=int)
    if emp:
        q = q.filter(StockOutput.employee_id == emp)
    reason = request.args.get("reason")
    if reason:
        q = q.filter(StockOutput.reason == reason)
    date_from = request.args.get("date_from")
    if date_from:
        try:
            q = q.filter(StockOutput.output_date >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    date_to = request.args.get("date_to")
    if date_to:
        try:
            q = q.filter(StockOutput.output_date <= datetime.fromisoformat(date_to))
        except ValueError:
            pass
    q = q.order_by(StockOutput.output_date.desc(), StockOutput.id.desc())
    result = paginate_query(q, default_per_page=25)
    return _ok({"items": [o.to_dict() for o in result["items"]],
                "meta": result["meta"]})


@bp.post("/outputs")
@require_permission("stock.output", api=True)
def outputs_create():
    data = request.get_json(silent=True) or {}
    pid = data.get("product_id")
    qty = _to_decimal(data.get("quantity"))
    if not pid:
        return _err("Produto é obrigatório.")
    if qty <= 0:
        return _err("Quantidade deve ser maior que zero.")
    product = Product.query.get(pid)
    if not product:
        return _err("Produto não encontrado.", 404)
    if (product.stock_quantity or 0) < qty:
        return _err(f"Estoque insuficiente. Disponível: {float(product.stock_quantity or 0)}",
                    status=400)

    unit_price = _to_decimal(data.get("unit_price", product.sale_price))
    out = StockOutput(
        company_id=_company_id(),
        product_id=product.id,
        employee_id=data.get("employee_id") or None,
        document=data.get("document"),
        quantity=qty,
        unit_price=unit_price,
        total_price=qty * unit_price,
        reason=data.get("reason") or "consumption",
        destination=data.get("destination"),
        notes=data.get("notes"),
        output_date=datetime.utcnow(),
        status="confirmed",
        created_by_id=current_user.id,
    )
    product.stock_quantity = (product.stock_quantity or 0) - qty
    db.session.add(out)
    db.session.flush()
    db.session.add(StockMovement(
        company_id=_company_id(),
        product_id=product.id,
        output_id=out.id,
        movement_type="output",
        document=out.document,
        quantity=-qty,
        unit_value=unit_price,
        total_value=qty * unit_price,
        balance_after=product.stock_quantity,
        reason=out.reason,
        movement_date=out.output_date,
        created_by_id=current_user.id,
    ))
    db.session.commit()

    # Low-stock notification
    if product.stock_quantity <= product.min_stock:
        n = Notification(
            company_id=_company_id(),
            type="warning",
            title="Estoque baixo",
            message=f"{product.sku} - {product.name} atingiu o estoque mínimo.",
            link=f"/app/produtos?id={product.id}",
        )
        db.session.add(n)
        db.session.commit()

    log_activity("output", "stock_output", out.id,
                 f"-{qty} de {product.sku} ({product.name})")
    return _ok(out.to_dict(), 201)


@bp.delete("/outputs/<int:oid>")
@require_permission("stock.output", api=True)
def outputs_delete(oid):
    o = StockOutput.query.get_or_404(oid)
    if o.status == "cancelled":
        return _err("Saída já cancelada.")
    p = o.product
    if p:
        p.stock_quantity = (p.stock_quantity or 0) + (o.quantity or 0)
    o.status = "cancelled"
    db.session.add(StockMovement(
        company_id=o.company_id,
        product_id=o.product_id,
        output_id=o.id,
        movement_type="cancel",
        document=o.document,
        quantity=o.quantity or 0,
        unit_value=o.unit_price,
        total_value=o.total_price or 0,
        balance_after=p.stock_quantity if p else 0,
        reason="Cancelamento de saída",
        movement_date=datetime.utcnow(),
        created_by_id=current_user.id,
    ))
    db.session.commit()
    log_activity("delete", "stock_output", o.id,
                 f"Saída cancelada (+{o.quantity} de {p.sku if p else '?'})")
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Stock snapshot
# ---------------------------------------------------------------------------
@bp.get("/stock")
@require_permission("stock.view", api=True)
def stock_snapshot():
    cid = _company_id()
    q = Product.query.filter_by(company_id=cid)
    status = request.args.get("stock_status")
    if status == "low":
        q = q.filter(Product.stock_quantity <= Product.min_stock,
                     Product.stock_quantity > 0)
    elif status == "out":
        q = q.filter(Product.stock_quantity <= 0)
    elif status == "ok":
        q = q.filter(Product.stock_quantity > Product.min_stock)
    search = (request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        q = q.filter(or_(Product.name.ilike(like), Product.sku.ilike(like)))
    q = q.order_by(Product.name.asc())
    result = paginate_query(q, default_per_page=50)
    return _ok({"items": [p.to_dict() for p in result["items"]],
                "meta": result["meta"]})


# ---------------------------------------------------------------------------
# Reports + Exports
# ---------------------------------------------------------------------------
def _gather_stock_rows():
    cid = _company_id()
    q = Product.query.filter_by(company_id=cid)
    return q.order_by(Product.name.asc()).all()


def _gather_movements(kind: str = "all", days: int = 30):
    cid = _company_id()
    since = datetime.utcnow() - timedelta(days=days)
    entries = outputs = []
    if kind in ("all", "entries"):
        eq = StockEntry.query.filter_by(company_id=cid)
        entries = eq.filter(StockEntry.entry_date >= since).all()
    if kind in ("all", "outputs"):
        oq = StockOutput.query.filter_by(company_id=cid)
        outputs = oq.filter(StockOutput.output_date >= since).all()
    return entries, outputs


@bp.get("/reports/summary")
@require_permission("reports.view", api=True)
def reports_summary():
    days = request.args.get("days", default=30, type=int)
    entries, outputs = _gather_movements("all", days)
    total_in = sum(float(e.quantity or 0) for e in entries)
    total_out = sum(float(o.quantity or 0) for o in outputs)
    value_in = sum(float(e.total_cost or 0) for e in entries)
    value_out = sum(float(o.total_price or 0) for o in outputs)
    return _ok({
        "period_days": days,
        "entries_count": len(entries),
        "outputs_count": len(outputs),
        "total_in": total_in,
        "total_out": total_out,
        "value_in": round(value_in, 2),
        "value_out": round(value_out, 2),
    })


@bp.get("/reports/export/stock.xlsx")
@require_permission("reports.export", api=True)
def export_stock_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    headers = ["SKU", "Produto", "Categoria", "Fornecedor", "Unidade",
               "Estoque Atual", "Mínimo", "Máximo", "Custo Unit.", "Valor Total", "Status"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = fill; c.font = font; c.alignment = Alignment(horizontal="center")

    for p in _gather_stock_rows():
        ws.append([
            p.sku, p.name,
            p.category.name if p.category else "",
            p.supplier.name if p.supplier else "",
            p.unit,
            float(p.stock_quantity or 0),
            float(p.min_stock or 0),
            float(p.max_stock or 0),
            float(p.cost_price or 0),
            float((p.stock_quantity or 0) * (p.cost_price or 0)),
            {"ok": "OK", "low": "BAIXO", "out": "ZERADO"}[p.stock_status],
        ])
    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(40, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    log_activity("export", "report", None, "Exportação Excel: Estoque")
    return send_file(buf, as_attachment=True,
                     download_name=f"movstok_estoque_{datetime.utcnow():%Y%m%d_%H%M}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.get("/reports/export/movements.xlsx")
@require_permission("reports.export", api=True)
def export_movements_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    days = request.args.get("days", default=30, type=int)
    entries, outputs = _gather_movements("all", days)

    wb = Workbook()
    ws1 = wb.active; ws1.title = "Entradas"
    for col, h in enumerate(["Data", "Documento", "SKU", "Produto", "Fornecedor",
                             "Qtde", "Custo Unit.", "Total", "Status"], 1):
        ws1.cell(row=1, column=col, value=h)
    for e in entries:
        ws1.append([
            e.entry_date.strftime("%d/%m/%Y %H:%M") if e.entry_date else "",
            e.document or "",
            e.product.sku if e.product else "",
            e.product.name if e.product else "",
            e.supplier.name if e.supplier else "",
            float(e.quantity or 0), float(e.unit_cost or 0), float(e.total_cost or 0),
            e.status,
        ])

    ws2 = wb.create_sheet("Saídas")
    for col, h in enumerate(["Data", "Documento", "SKU", "Produto", "Matrícula",
                             "Funcionário", "Motivo", "Qtde", "Preço Unit.", "Total", "Status"], 1):
        ws2.cell(row=1, column=col, value=h)
    for o in outputs:
        ws2.append([
            o.output_date.strftime("%d/%m/%Y %H:%M") if o.output_date else "",
            o.document or "",
            o.product.sku if o.product else "",
            o.product.name if o.product else "",
            o.employee.enrollment if o.employee else "",
            o.employee.name if o.employee else "",
            o.reason or "",
            float(o.quantity or 0), float(o.unit_price or 0), float(o.total_price or 0),
            o.status,
        ])

    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for ws in (ws1, ws2):
        for c in ws[1]:
            c.fill = fill; c.font = font; c.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(36, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    log_activity("export", "report", None, f"Exportação Excel: Movimentações ({days}d)")
    return send_file(buf, as_attachment=True,
                     download_name=f"movstok_movimentacoes_{datetime.utcnow():%Y%m%d_%H%M}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.get("/reports/export/stock.pdf")
@require_permission("reports.export", api=True)
def export_stock_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph("<b>MovStok - Relatório de Estoque</b>", styles["Title"])]
    story.append(Paragraph(
        f"Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["Normal"]))
    story.append(Spacer(1, 6 * mm))

    data = [["SKU", "Produto", "Categoria", "Un.", "Estoque",
             "Mín.", "Custo", "Valor Total", "Status"]]
    for p in _gather_stock_rows():
        data.append([
            p.sku, p.name[:40],
            p.category.name if p.category else "-",
            p.unit,
            f"{float(p.stock_quantity or 0):.0f}",
            f"{float(p.min_stock or 0):.0f}",
            f"R$ {float(p.cost_price or 0):.2f}",
            f"R$ {float((p.stock_quantity or 0) * (p.cost_price or 0)):.2f}",
            {"ok": "OK", "low": "BAIXO", "out": "ZERADO"}[p.stock_status],
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (-1, 1), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9DEE5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F3F5F7")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    doc.build(story)
    buf.seek(0)
    log_activity("export", "report", None, "Exportação PDF: Estoque")
    return send_file(buf, as_attachment=True,
                     download_name=f"movstok_estoque_{datetime.utcnow():%Y%m%d_%H%M}.pdf",
                     mimetype="application/pdf")


# ---------------------------------------------------------------------------
# Notifications
# ---------------------------------------------------------------------------
@bp.get("/notifications")
@login_required
def notifications_list():
    cid = _company_id()
    q = Notification.query.filter_by(company_id=cid)
    q = q.filter(or_(Notification.user_id == current_user.id,
                     Notification.user_id.is_(None)))
    q = q.order_by(Notification.created_at.desc()).limit(50)
    items = [n.to_dict() for n in q.all()]
    unread = sum(1 for n in items if not n["read"])
    return _ok({"items": items, "unread": unread})


@bp.post("/notifications/<int:nid>/read")
@login_required
def notifications_read(nid):
    n = Notification.query.get_or_404(nid)
    if not n.read_at:
        n.read_at = datetime.utcnow()
        db.session.commit()
    return _ok({"ok": True})


@bp.post("/notifications/read-all")
@login_required
def notifications_read_all():
    cid = _company_id()
    q = Notification.query.filter_by(company_id=cid).filter(Notification.read_at.is_(None))
    for n in q.all():
        n.read_at = datetime.utcnow()
    db.session.commit()
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Activity log
# ---------------------------------------------------------------------------
@bp.get("/activities")
@login_required
def activities_list():
    cid = _company_id()
    q = ActivityLog.query.filter_by(company_id=cid)
    action = request.args.get("action")
    if action:
        q = q.filter(ActivityLog.action == action)
    entity = request.args.get("entity")
    if entity:
        q = q.filter(ActivityLog.entity == entity)
    q = q.order_by(ActivityLog.created_at.desc())
    result = paginate_query(q, default_per_page=30)
    return _ok({"items": [a.to_dict() for a in result["items"]],
                "meta": result["meta"]})


# ---------------------------------------------------------------------------
# Company, Users, Roles
# ---------------------------------------------------------------------------
@bp.get("/company")
@login_required
def company_get():
    cid = _company_id()
    c = Company.query.get(cid)
    return _ok(c.to_dict() if c else {})


@bp.put("/company")
@require_permission("settings.manage", api=True)
def company_update():
    cid = _company_id()
    c = Company.query.get(cid)
    if not c:
        return _err("Empresa não encontrada.", 404)
    data = request.get_json(silent=True) or {}
    for f in ("name", "legal_name", "cnpj", "email", "phone",
              "address", "city", "state", "zipcode", "logo_url"):
        if f in data:
            setattr(c, f, data[f])
    db.session.commit()
    log_activity("update", "company", c.id, "Empresa atualizada")
    return _ok(c.to_dict())


@bp.get("/roles")
@login_required
def roles_list():
    return _ok({"items": [r.to_dict() for r in Role.query.order_by(Role.name).all()]})


@bp.get("/permissions")
@login_required
def permissions_list():
    return _ok({"items": [p.to_dict() for p in Permission.query.order_by(Permission.module, Permission.code).all()]})


@bp.get("/users")
@require_permission("users.manage", api=True)
def users_list():
    cid = _company_id()
    q = User.query.filter_by(company_id=cid).order_by(User.name.asc())
    search = (request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        q = q.filter(or_(User.name.ilike(like), User.email.ilike(like)))
    return _ok({"items": [u.to_dict() for u in q.all()]})


@bp.post("/users")
@require_permission("users.manage", api=True)
def users_create():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    if not name or not email or not password:
        return _err("Nome, email e senha são obrigatórios.")
    if len(password) < 6:
        return _err("Senha deve ter pelo menos 6 caracteres.")
    if User.query.filter_by(email=email).first():
        return _err("Email já cadastrado.")
    u = User(
        name=name, email=email,
        company_id=_company_id(),
        role_id=data.get("role_id") or None,
        phone=data.get("phone"),
        status=data.get("status") or "active",
    )
    u.set_password(password)
    db.session.add(u)
    db.session.commit()
    log_activity("create", "user", u.id, f"Usuário criado: {u.email}")
    return _ok(u.to_dict(), 201)


@bp.put("/users/<int:uid>")
@bp.patch("/users/<int:uid>")
@require_permission("users.manage", api=True)
def users_update(uid):
    u = User.query.get_or_404(uid)
    data = request.get_json(silent=True) or {}
    for f in ("name", "phone", "status", "avatar_url"):
        if f in data:
            setattr(u, f, data[f])
    if "role_id" in data:
        u.role_id = data["role_id"] or None
    if "password" in data and data["password"]:
        if len(data["password"]) < 6:
            return _err("Senha deve ter pelo menos 6 caracteres.")
        u.set_password(data["password"])
    db.session.commit()
    log_activity("update", "user", u.id, f"Usuário atualizado: {u.email}")
    return _ok(u.to_dict())


@bp.delete("/users/<int:uid>")
@require_permission("users.manage", api=True)
def users_delete(uid):
    if uid == current_user.id:
        return _err("Você não pode desativar a si mesmo.")
    u = User.query.get_or_404(uid)
    u.status = "inactive"
    db.session.commit()
    log_activity("delete", "user", uid, f"Usuário desativado: {u.email}")
    return _ok({"ok": True})


# ---------------------------------------------------------------------------
# Stock locations
# ---------------------------------------------------------------------------
@bp.get("/locations")
@login_required
def locations_list():
    cid = _company_id()
    q = StockLocation.query.filter_by(company_id=cid)
    return _ok({"items": [l.to_dict() for l in q.order_by(StockLocation.name).all()]})


@bp.post("/locations")
@require_permission("settings.manage", api=True)
def locations_create():
    data = request.get_json(silent=True) or {}
    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    if not code or not name:
        return _err("Código e nome são obrigatórios.")
    l = StockLocation(company_id=_company_id(), code=code, name=name,
                      description=data.get("description"))
    db.session.add(l)
    db.session.commit()
    return _ok(l.to_dict(), 201)


# ---------------------------------------------------------------------------
# Legacy endpoints (backwards compatibility with the old frontend/API)
# ---------------------------------------------------------------------------
@bp.get("/legacy/resumo")
@login_required
def legacy_resumo():
    cid = _company_id()
    total_products = Product.query.filter_by(company_id=cid).count()
    total_units = float(db.session.query(func.coalesce(func.sum(Product.stock_quantity), 0))
                        .filter(Product.company_id == cid).scalar() or 0)
    entries = StockEntry.query.filter_by(company_id=cid).count()
    outputs = StockOutput.query.filter_by(company_id=cid).count()
    low = Product.query.filter_by(company_id=cid).filter(Product.stock_quantity <= Product.min_stock).count()
    emps = Employee.query.filter_by(company_id=cid, status="active").count()
    return _ok({
        "total_produtos": total_products,
        "total_itens": int(total_units),
        "entradas": entries,
        "saidas": outputs,
        "baixo_estoque": low,
        "total_funcionarios": emps,
    })
